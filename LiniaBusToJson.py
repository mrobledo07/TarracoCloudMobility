import json
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook

def generar_json_horari_any(nom_excel, nom_full="Sheet1"):
    """
    Llegeix un Excel on la primera fila conté els noms de les parades
    (cada columna és una parada) i les files següents (a partir de la 2)
    són busos, on cada cel·la conté l'hora de pas del bus per la parada.
    
    Després, replica aquest horari per tots els dies de l'any 2025,
    assignant un valor d'ocupació aleatòria per a cada parada.
    """
    # Carreguem l'Excel
    wb = load_workbook(nom_excel, data_only=True)
    sheet = wb[nom_full]

    # La primera fila són els noms de les parades
    # Els llegim per tenir la llista de parades en ordre
    stops = []
    for cell in sheet[1]:
        if cell.value is not None:
            stops.append(str(cell.value))
        else:
            stops.append("")

    # A partir de la fila 2, cada fila és un bus
    # Guardem en memòria els horaris de cada bus (fila)
    bus_schedules = []
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        bus_id = f"bus{idx:03d}"  # bus001, bus002, etc.
        # Creem la llista de parades-hora per aquest bus
        bus_stops = []
        for col_index, hora in enumerate(row):
            # Si la cel·la té hora, la convertim a string
            hora_str = str(hora) if hora is not None else ""
            # L'ocupació encara no la posem aquí, la posarem quan repliquem el dia
            bus_stops.append({
                "parada": stops[col_index],
                "hora": hora_str
            })
        bus_schedules.append((bus_id, bus_stops))

    # Preparem l'estructura base per al JSON
    dades = {"linea_41": {}}

    # Definim el rang de dates (01/01/2025 - 31/12/2025)
    data_inici = datetime(2025, 1, 1)
    data_final = datetime(2025, 12, 31)

    data_actual = data_inici
    while data_actual <= data_final:
        # Format de data DD/MM
        data_str = data_actual.strftime("%d/%m")
        dades["linea_41"][data_str] = {}

        # Recorrem tots els busos i assignem ocupacions
        for (bus_id, bus_stops) in bus_schedules:
            # Creem una nova llista per a cada dia,
            # afegint l'ocupació aleatòria en cada parada
            parades_dia = []
            for p in bus_stops:
                parades_dia.append({
                    "parada": p["parada"],
                    "hora": p["hora"],
                    "ocupacio": random.randint(20, 80)  # o el rang que vulguis
                })
            # Assignem la llista de parades al bus corresponent
            dades["linea_41"][data_str][bus_id] = parades_dia

        # Passem al següent dia
        data_actual += timedelta(days=1)

    return dades

if __name__ == "__main__":
    # Nom de l'Excel i del full on es troba l'horari diari
    nom_excel = "Linea54.xlsx"
    nom_full = "Hoja1"  # o el nom que tingui el teu full

    # Generem el diccionari a partir de l'Excel
    dades = generar_json_horari_any(nom_excel, nom_full)

    # Guardem el resultat a un fitxer JSON
    nom_sortida = "linia54_any_complet.json"
    with open(nom_sortida, "w", encoding="utf-8") as f:
        json.dump(dades, f, indent=2, ensure_ascii=False)

    print(f"Fitxer JSON creat correctament: {nom_sortida}")