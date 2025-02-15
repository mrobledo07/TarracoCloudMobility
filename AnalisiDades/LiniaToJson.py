import json
import math
from datetime import datetime, timedelta
from openpyxl import load_workbook
import sys  

def seasonal_factor(month):
    """
    Retorna un factor multiplicador de ocupación según el mes.
    Por ejemplo, en diciembre (mes 12) se incrementa la ocupación y en verano (meses 7 y 8) se reduce.
    """
    factors = {
        1: 1.1,
        2: 0.6,
        3: 0.5,
        4: 0.5,
        5: 0.6,
        6: 1.7,
        7: 1.8,
        8: 1.9,
        9: 1.0,
        10: 1.0,
        11: 0.9,
        12: 1.2
    }
    return factors.get(month, 1.0)

def occupancy_from_time(time_str, month, peak_time=13.0, sigma=7.0, min_occ=5, max_occ=100, stop_factor=1.0):
    """
    Calcula la ocupación basándose en la hora con una función gaussiana,
    y la ajusta con un factor estacional (según el mes) y un factor propio de la parada.
    
    Parámetros:
      - time_str: Cadena con la hora (formato "HH:MM" o "HH:MM:SS").
      - month: Mes (entero 1-12) para aplicar el factor estacional.
      - peak_time: Hora de pico para esta parada.
      - sigma: Desviación típica que controla el ancho de la campana.
      - min_occ: Ocupación mínima.
      - max_occ: Ocupación máxima.
      - stop_factor: Factor multiplicador propio de la parada.
    """
    if not time_str or time_str.strip() == "":
        return min_occ
    try:
        parts = time_str.split(":")
        hour = float(parts[0])
        minute = float(parts[1]) if len(parts) > 1 else 0
        t = hour + minute / 60.0
    except Exception:
        return min_occ

    # Cálculo de la distribución gaussiana
    gauss = math.exp(-((t - peak_time) ** 2) / (2 * sigma ** 2))
    occ = min_occ + (max_occ - min_occ) * gauss

    # Se ajusta la ocupación según la estacionalidad y la especificidad de la parada
    occ = occ * seasonal_factor(month) * stop_factor

    # Se limita el resultado al rango [min_occ, max_occ]
    occ = max(min_occ, min(max_occ, occ))
    return int(round(occ))

def generar_json_horari_any(nom_excel, line_number, nom_full="Sheet1"):
    """
    Lee un Excel en el que la primera fila contiene los nombres de las paradas (cada columna es una parada)
    y las filas siguientes (a partir de la 2) son buses, donde cada celda contiene la hora de paso del bus por la parada.
    
    Se replica este horario para cada día del año 2025, asignando a cada parada un valor de ocupación que depende de:
      - La hora (usando una función gaussiana).
      - El mes (ajuste estacional).
      - Parámetros específicos de cada parada (peak_time, sigma y stop_factor).
    """
    # Cargamos el Excel
    wb = load_workbook(nom_excel, data_only=True)
    sheet = wb[nom_full]

    # Se extraen los nombres de las paradas (la primera fila)
    stops = []
    for cell in sheet[1]:
        stops.append(str(cell.value) if cell.value is not None else "")

    # Se obtiene el horario de cada bus (cada fila a partir de la 2)
    bus_schedules = []
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        bus_id = f"bus{idx:03d}"
        bus_stops = []
        for col_index, hora in enumerate(row):
            hora_str = str(hora) if hora is not None else ""
            bus_stops.append({
                "parada": stops[col_index],
                "hora": hora_str
            })
        bus_schedules.append((bus_id, bus_stops))

    # Definir parámetros específicos por parada.
    # Puedes ajustar estos valores según la popularidad o características de cada parada.
    # Definir parámetros específicos por parada.
    # Definir parámetros específicos por parada.
    stop_parameters = {
        # Paradas principales - Alta demanda (centro ciudad/transporte)
        "Estació": {"peak_time": 12.8, "sigma": 6.8, "stop_factor": 1.25},
        "Pl. Imperial Tàrraco": {"peak_time": 13.3, "sigma": 7.3, "stop_factor": 1.35},
        
        # Zonas comerciales - Fluctuación media-alta
        "Pl. Carros": {"peak_time": 13.2, "sigma": 7.2, "stop_factor": 1.15},
        "Reial": {"peak_time": 12.7, "sigma": 6.9, "stop_factor": 0.85},
        "Pere Martell, 54": {"peak_time": 13.0, "sigma": 6.7, "stop_factor": 1.25},
        "Pere Martell, 34": {"peak_time": 13.2, "sigma": 7.1, "stop_factor": 0.95},
        
        # Zonas residenciales - Fluctuación baja-media
        "Claret": {"peak_time": 12.4, "sigma": 6.4, "stop_factor": 0.65},
        "Montoliu": {"peak_time": 13.5, "sigma": 7.5, "stop_factor": 0.78},
        "Torreforta": {"peak_time": 12.6, "sigma": 6.7, "stop_factor": 0.72},
        
        # Zonas educativas - Alta variación según horario
        "4 Garrofers": {"peak_time": 13.0, "sigma": 7.0, "stop_factor": 1.45},
        "Educacional N-240": {"peak_time": 12.7, "sigma": 6.9, "stop_factor": 0.88},
        "Escola d'art": {"peak_time": 13.4, "sigma": 7.4, "stop_factor": 1.25},
        "Politècnic": {"peak_time": 12.6, "sigma": 6.8, "stop_factor": 1.65},  # Máximo pico
        
        # Paradas periféricas - Mayor fluctuación
        "Bonavista": {"peak_time": 12.5, "sigma": 6.5, "stop_factor": 0.55},
        "Carrer Vint-i-tres": {"peak_time": 13.2, "sigma": 7.1, "stop_factor": 0.42},
        "Mercadet": {"peak_time": 12.8, "sigma": 6.7, "stop_factor": 1.18},
        
        # Paradas especiales (hospitales, centros administrativos)
        "La Salle": {"peak_time": 12.9, "sigma": 6.6, "stop_factor": 1.32},
        "St. Benilde": {"peak_time": 13.1, "sigma": 7.2, "stop_factor": 0.68},
        
        # Ajustes extremos para testear variabilidad
        "Països Catalans, 41": {"peak_time": 13.3, "sigma": 7.3, "stop_factor": 1.82},
        "Pl. Generalitat": {"peak_time": 12.4, "sigma": 6.4, "stop_factor": 0.38},
        
        # ... (mantenemos el resto de paradas con ajustes similares)
        "Jaume I": {"peak_time": 13.0, "sigma": 7.0, "stop_factor": 1.08},
        "Prat de la Riba": {"peak_time": 12.9, "sigma": 6.6, "stop_factor": 0.92},
        "Rambla Nova, 105": {"peak_time": 13.3, "sigma": 7.3, "stop_factor": 1.15},
        "Guatemala": {"peak_time": 13.4, "sigma": 7.4, "stop_factor": 0.79},
        "Coop. Tàrraco": {"peak_time": 13.2, "sigma": 7.2, "stop_factor": 1.27}
    }


    # Clave dinámica para la línea
    linea_key = f"linia_{line_number}"
    dades = {linea_key: {}}

    data_inici = datetime(2025, 1, 1)
    data_final = datetime(2025, 12, 31)
    data_actual = data_inici

    while data_actual <= data_final:
        data_str = data_actual.strftime("%d/%m")
        dades[linea_key][data_str] = {}
        month = data_actual.month

        for (bus_id, bus_stops) in bus_schedules:
            parades_dia = []
            for p in bus_stops:
                # Se obtienen los parámetros específicos para la parada o se usan valores por defecto
                params = stop_parameters.get(p["parada"], {"peak_time": 13.0, "sigma": 7.0, "stop_factor": 1.0})
                occ = occupancy_from_time(p["hora"], month,
                                          peak_time=params["peak_time"],
                                          sigma=params["sigma"],
                                          stop_factor=params["stop_factor"])
                parades_dia.append({
                    "parada": p["parada"],
                    "hora": p["hora"],
                    "ocupacio": occ
                })
            dades[linea_key][data_str][bus_id] = parades_dia

        data_actual += timedelta(days=1)

    return dades

if __name__ == "__main__":
    # Validar parámetro de línea
    if len(sys.argv) != 2:
        print("Uso: python script.py <número_línea>")
        print("Ejemplo: python script.py 41")
        exit(1)
    
    line_number = sys.argv[1]
    if line_number not in ("41", "54"):
        print("Error: Número de línea debe ser 41 o 54")
        exit(1)
    
    # Nombres de archivos dinámicos
    nom_excel = f"Linia{line_number}.xlsx"
    nom_sortida = f"linia{line_number}_any_complet.json"
    
    # Generar datos
    dades = generar_json_horari_any(nom_excel, line_number, nom_full="Hoja1")
    
    # Guardar resultado
    with open(nom_sortida, "w", encoding="utf-8") as f:
        json.dump(dades, f, indent=2, ensure_ascii=False)
    
    print(f"Archivo JSON creado: {nom_sortida}")
